import openpyxl

path = "D:/SeleniumPython/testdatadriven.xlsx"

workbook = openpyxl.load_workbook(path)
#sheet = workbook.get_sheet_by_name("sheet1")
sheet = workbook.active
row = sheet.max_row
col = sheet.max_column
print(row,col)

for r in range(1,row+1):
    for c in range(1,col+1):
        print(sheet.cell(row=r,column=c).value, end = " ")
    print(" ")
        